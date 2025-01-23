# Running in python 3.10
from lingpy import *
#from pandas import read_excel
import openpyxl
import xlsxwriter
import re

# Local file
from CLTSFeatureBasedAlignment import *

normalisePartialWordComparisonForLength = False

# TODO: (1) are the measures coming out of the analysis similarity or difference?
#  (2) are the measures being normalised properly?

# Nerbonne & Heeringa (2010) have a gap cost of 1
simple_gap_cost = -1 
feature_gap_cost = 0.1*-0.4338
historical_gap_cost = 0.1*-0.4609

vowels = ['V', 'a', 'a:', 'aɪ', 'aʊ', 'e', 'e:', 'eɪ', 'eʊ', 'i:', 'o', 'o:', 'u:', 'æ:', 'ø', 'øʊ', 'ɔ', 'ɔ:', 'ɔʊ', 'ə', 'ɛ', 'ɛ:', 'ɛ:ʊ', 'ɛɪ', 'ɛʊ', 'ɪ', 'ɪ:', 'ʊ']

slashToUncertaintySymbols = {}
uncertaintySymbolsToSlash = {}

def cleanWord(word):
	word = str(word)
	# Remove anything between parentheses
	#if word.count("(") >0 and word.count(")")>0:
	#	word = re.sub("\\(.+?\\)","",word)
	#if word.count(" or ")>0:
	#	word = word[:word.index(" or ")].strip()
	word = word.replace("8","ə")
	word = word.replace("Ə","ə")
	word = word.replace("0","ɔ")
	word = word.replace("3","ɛ")
	word = word.replace("th","θ")
	word = word.replace("dh","ð")
	word = word.replace("sh","ʃ")
	word = word.replace("ch","tʃ") # affricate 
	word = word.replace("mm","m m")
	word = word.replace("þ","Þ")
	word = word.replace("ː",":")	
	#word = word.replace("?","s")	
	word = word.replace("ɡ","g")
	
	# W should be treated as IPA [w]
	word = word.replace("W","w")
	
	word = re.sub("\\[.+?\\]","",word)
	word = word.replace("-","")
	word = word.strip()
	
	return(word)

def KScore(a,b):
	# Score between two segments
	# From Keller (2020)
	# Insertions and deletions are worth one point of distance, 
	#  as are substitutions of a vowel with a consonant or vice versa. 
	# Substitutions of vowels with vowels or consonants with consonants, 
	#  however, are only awarded a distance of 0.5. 
	#[Vowel] Length is understood to be part of the preceding sound 
	#  and adds a distance of 0.25. 
	
	# From Gooskens 2017:
	#In the weighing of the differences between two letters, a distinction was made between differences in the base of the letter and differences in diacritics. Comparing Danish hånd with English hand, for example, we see that the å in the Danish form contains a diacritic, while the a in the English form does not. Differences like this were weighed 0.3. Differences in the base of the character, like the difference between German Pro z ent and Dutch pro c ent ‘percent’ were weighed 1. The maximum weighing was 1, so if two characters differed in both base and diacritics (e.g. o versus å), they got a weighing of 1 and not of 1.3.
	
	# From 	Gooskens, C., & van Heuven, V. (2020). How well can intelligibility of closely related languages in Europe be predicted by linguistic and non-linguistic variables? Linguistic Approaches to Bilingualism, 10(3), 351-379. https://doi.org/10.1075/lab.17084.goo
	# To constrain possible alignments, vowels match with vowels and consonants with consonants but [j, w] also with vowels and schwa with sonorants.
	
	# Both point back to Nerbonne & Heeringa (2010)
	# From https://www.let.rug.nl/nerbonne/papers/HSK-Nerbonne-Web-Version2.pdf " table of segment distances, however derived (called alphabetic weights in Gusfield, 1999), "
	# Nerbonne, J., & Heeringa, W. (2010). Measuring dialect differences. Language and Space: Theories and Methods. Berlin: Mouton De Gruyter, 550-566.
	# " for example, the confusability matrices phoneticians compile from how frequently one segment may be mistaken for another (Johnson, 2004, Chap.4)"
	
	# Treating diphthongs as simple vowels

	segmentsAreIdentical = False
	if a==b:
		segmentsAreIdentical = True
	elif (a=="V" and (b in vowels)) or (b=="V" and (a in vowels)):
		# General vowel cateogry
		segmentsAreIdentical = True
	elif a == "F":
		# Voiced or voiceless labiodental fricative
		segmentsAreIdentical = b in ["F","f","v"]
	elif b == "F":
		segmentsAreIdentical = a in ["F","f","v"]
	elif a == "Θ":
		# Voiced or voiceless dental fricative
		segmentsAreIdentical = b in ["Θ","θ","ð"]
	elif b == "Θ":
		segmentsAreIdentical = a in ["Θ","θ","ð"]


	score = 0
	if segmentsAreIdentical:
		score = 0 # Segments are the same
	else:
		if a.replace(":","") in vowels and b.replace(":","") in vowels:
			if a.count(":")>0 or b.count(":")>0:
				if a.replace(":","") == b.replace(":",""):
					score = 0.25 # Same vowel, but different length 
				else:
					score = 0.75 # different vowels and different lengths (0.5 + 0.25)
			else:
				score = 0.5 # Vowel with vowel
		elif (not a in vowels) and (not b in vowels):
			score = 0.5 # consonant with consonant
		else:
			score = 1 # vowel with consonant
			
	# The alignment algorithm requires similarity, not distance,
	#  so need to flip to similarity score
	score = 1-score
	
	return(score)

def getSimpleScorer(tokens):
	scorer = {}
	for t1 in tokens:
		for t2 in tokens:
			ks = KScore(t1,t2)
			scorer[(t1,t2)] = ks
	return(scorer)

# def getSimpleScorer2(seqA,seqB):
# 	
# 	#print((seqA,seqB))
# 	tokA = tokenise(seqA)
# 	tokB = tokenise(seqB)
# 	#print((tokA,tokB))
# 	scorer={}
# 	# Score is alignment value, not distance
# 	for a in tokA:
# 		for b in tokB:
# 			if a is str and b is str:
# 				scorer[(a,b)] = KScore(a,b)
# # 			elif "ðə" in a or "ðə" in b:
# # 				# For the special case of "baː(θ)/(ðə)" i.e. either baːθ or baːðə
# # 				aSequences = ["".join([sel(x,i) for x in tokA]) for i in range(max([len(z) for z in tokA]))]
# # 				bSequences = ["".join([sel(x,i) for x in tokB]) for i in range(max([len(z) for z in tokB]))]
# 			else:
# 				# At least one of the forms is a tuple
# 				partScores = []
# 				for aParts in a:
# 					for bParts in b:
# 						partScores.append(KScore(aParts,bParts))
# 				scorer[(a,b)] = max(partScores)
# 	return(scorer)
	

def tokenise(seq):
	if type(seq) == list:
		return(seq) # Already tokenised
	
	if seq == "(ɔ)/(ɔ:)l(ə)/()":
		# There's one case with two uncertainties, so manually fix:
		return([("ɔ","ɔ:"),"l",("ə","")])
	
	toks = ""
	if seq.count("/")>0:
		# Deal with uncertainty in some segments - add as a tuple of possible segments
		pre = seq[:seq.index("(")]
		mid = seq[seq.index("("):seq.rindex(")")+1]
		post = seq[seq.rindex(")")+1:]
	
		toks_pre = []
		if len(pre)>0:
			toks_pre = ipa2tokens(pre.replace(" ",""))
		possibleSegments = tuple([x.replace("(","").replace(")","") for x in mid.split("/")])
		toks_post = []
		if len(post)>0:
			toks_post = ipa2tokens(post.replace(" ",""))
		toks = toks_pre + [possibleSegments] + toks_post
	else:
		toks = ipa2tokens(seq.replace(" ",""))
	out = []
	for tok in toks:
		if tok == "mm":
			out += ['m','m']
		elif tok is str and tok.startswith("y") and len(tok)>1:
			out += ["y",tok[1:]]
		else:
			out.append(tok)
	
	return(out)


def sel(x,i):
	if type(x) is tuple and len(x)>1:
		return(x[i])
	else:
		return(x)
		
def tupleLen(x):
	if type(x) is tuple:
		return(len(x))
	else:
		return(1)	

def NWAlign(seqA,seqB,scorer,gap_cost):
	tokA = tokenise(seqA)
	tokB = tokenise(seqB)
	bestResult = ([],[],10000,10000)
	if any([type(x) is tuple for x in tokA]) or any([type(x) is tuple for x in tokB]):
		toksA = [[sel(x,i) for x in tokA] for i in range(max([tupleLen(z) for z in tokA]))]
		toksB = [[sel(x,i) for x in tokB] for i in range(max([tupleLen(z) for z in tokB]))]
		totalScore = 0
		numComparisons = 0
		for ta in toksA:
			for tb in toksB:
				# Need to re-tokenise, because some alt definitions have multiple segments
				ta = tokenise("".join(ta))
				tb = tokenise("".join(tb))
				ax = NWAlign2(ta,tb,scorer,gap_cost)
				if ax[2] < bestResult[2]:
					bestResult = ax
		return(bestResult)
	else:
		return(NWAlign2(tokA,tokB,scorer,gap_cost))
#	return(NWAlign(tokA,tokB,scorer,gap_cost))

def NWAlign2(tokA,tokB,scorer,gap_cost):
	# However, NW works by alignment, not distance, so need to flip
	# see also pw_align, which has a distance option
	ax = align.pairwise.nw_align(tokA,tokB,scorer = scorer,gap = gap_cost)
	# Flip alignment to negative to be distance
	# TODO: should this be the number of alignments, not the length of the longest string?
	normSim = (-ax[2])/max([len(ax[0]),len(ax[1])])
	# Align A, Align B, total dist, normed dist
	return((ax[0],ax[1],-ax[2],normSim))
	
def getAlliteration(seqA,seqB):
	seqA = tokenise(seqA)
	seqB = tokenise(seqB)

	#Consonant clusters that have to be considered: 
	#s, sk, sp, st and sh are normally treated as distinct, so each one alliterates only with itself.	
	# if both sequences start with 's' ...
	if seqA[0]=="s" and seqB[0]=="s":
		# ... we have to check the second segment
		if seqA[1] in ["k","p","t","h"] or seqB[1] in ["k","p","t","h"]:
			# if they're both cluster types, return whether the second segments are different.
			return(seqA[1]!=seqB[1])
		else:
			# otherwise, a plain 's', sequences alliterate, no advantage in choosing either
			return(False)
	# else, check whether both start with vowels or 'h', 
	elif seqA[0] in vowels+['h'] and seqB[0] in vowels+['h']:
		# A vowel/diphthong alliterates with any other vowel/diphthong, 
		# so a poet would not win anything by using e.g. aue vs eie in that respect. 
		# h can alliterate with h and with any vowel
		return(False)
	else:
		# Otherwise, just check whether the first segments are different
		return(seqA[0]!=seqB[0])
	


d = openpyxl.load_workbook("../data/SharedIntegrationOfCognatesData.xlsx")
dataframe1 = d.worksheets[0]

headers = [col[0].value for col in dataframe1.iter_cols(0, dataframe1.max_column)]

print(headers)

data = {}
wordClasses = {}
# Iterate the loop to read the cell values
idnum = 0
for row in range(1, dataframe1.max_row):
	idnum += 1
	vals = [col[row].value for col in dataframe1.iter_cols(1, dataframe1.max_column)]
	setNum = str(vals[headers.index("Set")]).strip()
	# Numerals in the Set name should be ignored:
	setNum = re.sub("[0-9]","",setNum)
	lex = vals[headers.index("Lexeme")]
	etym = vals[headers.index("Etymology")]
	if etym!="?":
		wclass = vals[headers.index("Class")]
		if not wclass is None:
			wordClasses[setNum] = wclass
		else:
			if setNum in wordClasses:
				wclass = wordClasses[setNum]
		freqs = vals[headers.index("No. in Ormulum"):(1+headers.index("Rolle"))]

		formForComp = vals[headers.index("Diagnostic features")]
		word = vals[headers.index("FullForm")]
		if not word is None:
			word = cleanWord(word)
			if not setNum in data:
				data[setNum] = []
			data[setNum].append({"word":word, "class":wclass, "etym":etym, "freqs":freqs, "lex":lex, "formForComp": formForComp, "origWord": vals[headers.index("Attested form")]})
		



# Make the feature scorer
#  by finding all the unique segments
allTokens = []
for setNum in data:
	for row in data[setNum]:
		word = row["word"]
		tokens = tokenise(word)
		for token in tokens:
			if type(token) is tuple:
				for tx in token:
					if len(tx.strip())>0:
						txx = tokenise(tx) # Some alternatives have multiple segments
						allTokens += txx
			else:
				allTokens.append(token)
allTokens = list(set(allTokens))
allTokens = [x for x in allTokens if len(x)>0 and not x in ["ew","ðə","ɛ:wə","øw","aw","ɛj"]]
		
print(allTokens)
print([x for x in allTokens if not x in vowels])

# Get all relevant features for each class that are not None
features = {}
featureValues = {}
for token in allTokens:
	print(token)
	x = getFeatures(token)
	presentFeatures = [k for k in x.featuredict if not x.featuredict[k] is None]
	if not x.type in features:
		features[x.type] = []
	features[x.type] = list(set(features[x.type] + presentFeatures))
	for k in presentFeatures:
		if not k in featureValues:
			featureValues[k] = []
		featureValues[k] = list(set(featureValues[k]+[x.featuredict[k]]))

print(features)
print(featureValues)

simpleScorer = getSimpleScorer(allTokens)

# Function from CLTSFeatureBasedAlignment.py
featureScorer = get_feature_scorer(allTokens,features=features)

# Make a scorer based on Jager (2015)
historicalScorer = get_historical_scorer(allTokens)


seqA = "hʊndərθ"#hʊndVrd
seqB = "hʊndrəd"
#simpleScorer = getSimpleScorer(seqA,seqB)
print("\n".join([str(x) for x in NWAlign(seqA,seqB,simpleScorer,simple_gap_cost)]))
print("\n".join([str(x) for x in NWAlign(seqA,seqB,featureScorer,feature_gap_cost)]))
print("\n".join([str(x) for x in NWAlign(seqA,seqB,historicalScorer,historical_gap_cost)]))

print('------')

seqA = "grɛ:θ"
seqB = "rɛ:d"
#simpleScorer = getSimpleScorer(seqA,seqB)
print("\n".join([str(x) for x in NWAlign(seqA,seqB,simpleScorer,simple_gap_cost)]))
print("\n".join([str(x) for x in NWAlign(seqA,seqB,featureScorer,feature_gap_cost)]))
print("\n".join([str(x) for x in NWAlign(seqA,seqB,historicalScorer,historical_gap_cost)]))

fst = [featureScorer[(a,b)] for a,b in featureScorer if a!=b]
print(sum(fst)/len(fst))
hst = [historicalScorer[(a,b)] for a,b in historicalScorer if a!=b]
print(sum(hst)/len(hst))

print("TEST1")
print("\n".join([str(x) for x in NWAlign("bkdfg","bkfg",simpleScorer,simple_gap_cost)]))


outputFilename = "../data/IntegrationDistances.xlsx"
workbook = xlsxwriter.Workbook(outputFilename)
worksheet = workbook.add_worksheet("Distances")
worksheet.set_default_row(30)
text_format = workbook.add_format({'text_wrap': True, 'font_name':"Courier New"})
worksheet.set_column(2, 13, 30)

xlsxHeaders = ["Set",
	"Class",
	"NorseLexeme",
	"EngLexeme",
	"NorseForm",
	"EngForm",
	"NorseFormDiagnostic",
	"EngFormDiagnostic",
	# Dists
	"Alignment",
	"RawDistance",
	"NormDistance",
	"FeatureAlignment",
	"RawFeatureDistance",
	"NormFeatureDistance",
	"HistoricalAlignment",
	"RawHistoricalDistance",
	"NormHistoricalDistance",
	# N Freqs
	"NFreqOrmulum",
	"NFreqFCPC",
	"NFreqHavelok",
	"NFreqGenAndEx",
	"NFreqMannyng",
	"NFreqGawainPoet",
	"NFreqWarsAlexander",
	"NFreqStErkenwald",
	"NFreqCursorMundi",
	"NFreqLinc",
	"NFreqNott",
	"NFreqNorf",
	"NFreqRolle",
	# E Freqs
	"EFreqOrmulum",
	"EFreqFCPC",
	"EFreqHavelok",
	"EFreqGenAndEx",
	"EFreqMannyng",
	"EFreqGawainPoet",
	"EFreqWarsAlexander",
	"EFreqStErkenwald",
	"EFreqCursorMundi",
	"EFreqLinc",
	"EFreqNott",
	"EFreqNorf",
	"EFreqRolle",
	# 
	"Alliteration"
	]


for i,h in enumerate(xlsxHeaders):
	worksheet.write(0,i,h)


rowNum = 1
for setNum in data:
	for NorseDat in [x for x in data[setNum] if x["etym"]=="Norse"]:
		NorseWord = NorseDat["word"]
		NorseWordFC = NorseDat["formForComp"]
		for EngDat in [x for x in data[setNum] if x["etym"]=="English"]:
			EngWord = EngDat["word"]
			EngWordFC = EngDat["formForComp"]
			wclass = EngDat["class"]
			
			# Simple score
			#simpleScorer = getSimpleScorer(NorseWord,EngWord)
			nwa = NWAlign(NorseWord,EngWord,simpleScorer,simple_gap_cost)
			simpleScoreTotal = nwa[2]
			simpleScoreNormed = nwa[3]
			
			# Feature-based score
			nwaFeature = NWAlign(NorseWord,EngWord,featureScorer,feature_gap_cost)
			featureScoreTotal = nwaFeature[2]
			featureScoreNormed = nwaFeature[3]
			
			# Historical scorer
			nwaHist = NWAlign(NorseWord,EngWord,historicalScorer,historical_gap_cost)
			historicalScoreTotal = nwaHist[2]
			historicalScoreNormed = nwaHist[3]
			
			# Normalise for other parts of the word that are not compared using alignment.
			EngOrigWord = cleanWord(EngDat["origWord"])
			# (turned off above)
			if normalisePartialWordComparisonForLength:
				if len(EngWord)< len(EngOrigWord):
					# assuming the scorer returns 0 for each segment in root
					# e.g. for one extra segment, add 0 and divide by 2
					simpleScoreNormed  = simpleScoreNormed / (1+ len(EngOrigWord) - len(EngWord) )
					featureScoreNormed = featureScoreNormed / (1+len(EngOrigWord) - len(EngWord) )
					historicalScoreNormed = historicalScoreNormed / (1+len(EngOrigWord) - len(EngWord) )
			
			#print((EngWord,NorseWord,NorseDat["freqs"]))
			#print(nwa[0])
			#print(nwa[1])
			#print((nwa[2],nwa[3]))
			
			# Alliteration
			alliteration = getAlliteration(NorseWord,EngWord)
			
			colNum = 0   
			worksheet.write(rowNum,colNum,setNum)
			colNum += 1
			worksheet.write(rowNum,colNum,wclass)
			colNum += 1
			worksheet.write(rowNum,colNum,NorseDat["lex"])
			colNum += 1
			worksheet.write(rowNum,colNum,EngDat["lex"])
			colNum += 1
			worksheet.write(rowNum,colNum,NorseWord)
			colNum += 1
			worksheet.write(rowNum,colNum,EngWord)
			colNum += 1
			worksheet.write(rowNum,colNum,NorseWordFC)
			colNum += 1
			worksheet.write(rowNum,colNum,EngWordFC)
			colNum += 1
			
			worksheet.write(rowNum,colNum," ".join(nwa[0])+'\n'+" ".join(nwa[1]),text_format)
			colNum += 1
			worksheet.write(rowNum,colNum,simpleScoreTotal)
			colNum += 1
			worksheet.write(rowNum,colNum,simpleScoreNormed)
			colNum += 1
			worksheet.write(rowNum,colNum," ".join(nwaFeature[0])+'\n'+" ".join(nwaFeature[1]),text_format)
			colNum += 1
			worksheet.write(rowNum,colNum,featureScoreTotal)
			colNum += 1
			worksheet.write(rowNum,colNum,featureScoreNormed)
			colNum += 1
			worksheet.write(rowNum,colNum," ".join(nwaHist[0])+'\n'+" ".join(nwaHist[1]),text_format)
			colNum += 1
			worksheet.write(rowNum,colNum,historicalScoreTotal)
			colNum += 1
			worksheet.write(rowNum,colNum,historicalScoreNormed)			
			colNum += 1
			
			for i in range(len(NorseDat["freqs"])):
				worksheet.write(rowNum,colNum,NorseDat["freqs"][i])
				colNum += 1
				
			for i in range(len(EngDat["freqs"])):
				worksheet.write(rowNum,colNum,EngDat["freqs"][i])
				colNum += 1
				
			worksheet.write(rowNum,colNum,alliteration)
			colNum += 1
			
			rowNum += 1

workbook.close()